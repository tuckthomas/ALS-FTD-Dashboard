import requests
import os
import re
import ast
from openai import OpenAI
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from django.conf import settings
from bs4 import BeautifulSoup
from django.db import models
from django.db.models import Q
import numpy as np
from .models import Trial, Gene, Update_Log, HealeyTrial, Intervention
from .schemas import HealeyTrialSchema, HealeyContactInfoSchema
from datetime import datetime, timedelta
from dateutil import parser as date_parser
from django.utils import timezone
from django.core.exceptions import ValidationError
from django.core.serializers.json import DjangoJSONEncoder
from fuzzywuzzy import process
import json
from typing import List
from django.http import JsonResponse
from django.db import transaction
import csv
from io import StringIO
import logging
import time

# Configure Logging
log_dir = os.path.join(settings.BASE_DIR, 'logs')
os.makedirs(log_dir, exist_ok=True)

# LLM Logger
llm_logger = logging.getLogger('llm_logger')
llm_logger.setLevel(logging.INFO)
file_handler = logging.FileHandler(os.path.join(log_dir, 'llm_classification.log'))
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)
llm_logger.addHandler(file_handler)


# This function is designed to update the database with trial and gene data.
# It iterates over each trial, updating or creating trial records, and associates genes by creating or fetching gene records.
# The logic for dynamically using field names from the Trial model minimizes hardcoding and adapts to changes in the model's structure.
def update_data():
    print("Inside update_data function...")
    trials_data = enhanced_fetch_trial_data()  # Ensure this returns a DataFrame
    print(f"Fetched {len(trials_data)} trial records (with gene matching).")

    gene_list_df = scrape_alsod_gene_list()  # Fetch the gene list
    print(f"Scraped gene list with {len(gene_list_df)} records.")
    valid_gene_symbols = set(gene_list_df['Gene Symbol'].tolist())

    # Update or create Gene records
    updated_genes = 0
    for _, row in gene_list_df.iterrows():
        _, created = Gene.objects.update_or_create(
            gene_symbol=row['Gene Symbol'],
            defaults={
                'gene_name': row['Gene Name'],
                'gene_risk_category': row['Gene Risk Category']
            }
        )
        updated_genes += (1 if created else 0)
    print(f"Total genes updated or created: {updated_genes}")

    updated_trials = 0
    updated_trial_ids = set()  # Collect IDs of trials being updated or created

    # Iterate through each trial record
    for _, row in trials_data.iterrows():
        trial_defaults = row.to_dict()

        # Convert date strings to date objects or None if invalid
        date_fields = ['study_submitance_date', 'study_submitance_date_qc', 'study_start_date', 'status_verified_date', 'completion_date']
        for date_field in date_fields:
            date_value = trial_defaults.get(date_field)
            if date_value:
                try:
                    # Use dateutil.parser to parse the date string
                    trial_defaults[date_field] = date_parser.parse(date_value).date()
                except (ValueError, TypeError):
                    print(f"Invalid date format for {date_field} in trial ID {trial_defaults['unique_protocol_id']}: {date_value}")
                    trial_defaults[date_field] = None
            else:
                # If the date is empty or None, explicitly set it to None
                trial_defaults[date_field] = None

        # Handling numerical fields with more robust error checking
        num_fields = ['enrollment_count']
        for num_field in num_fields:
            value = trial_defaults.get(num_field)
            if value is not None and value != '':
                try:
                    trial_defaults[num_field] = int(value)
                except ValueError:
                    # Log the error and set the field to None or a default value
                    print(f"Error converting {num_field} to int for trial ID {trial_defaults['unique_protocol_id']}: {value}")
                    trial_defaults[num_field] = None  # Or set a default value if appropriate
            else:
                # If the value is empty or None, explicitly set it to None
                trial_defaults[num_field] = None

        # Ensure JSON fields are properly parsed
        for json_field in ['genes', 'condition', 'intervention_name', 'keyword']:
            if trial_defaults.get(json_field) and isinstance(trial_defaults[json_field], str):
                try:
                    trial_defaults[json_field] = json.loads(trial_defaults[json_field])
                except json.JSONDecodeError:
                    print(f"Error parsing JSON for {json_field} in trial ID {trial_defaults['unique_protocol_id']}")
                    trial_defaults[json_field] = None

        # Update or create the Trial record
        trial_obj, created = Trial.objects.update_or_create(
            unique_protocol_id=trial_defaults['unique_protocol_id'],
            defaults=trial_defaults
        )
        updated_trials += (1 if created else 0)

        # Sync related_genes M2M field
        if trial_defaults.get('genes'):
            gene_symbols = trial_defaults['genes']
            if isinstance(gene_symbols, list):
                # match_genes returns a list of symbols (strings)
                genes_to_link = Gene.objects.filter(gene_symbol__in=gene_symbols)
                trial_obj.related_genes.set(genes_to_link)
            elif isinstance(gene_symbols, str):
                try:
                    gene_list = json.loads(gene_symbols)
                    if isinstance(gene_list, list):
                         genes_to_link = Gene.objects.filter(gene_symbol__in=gene_list)
                         trial_obj.related_genes.set(genes_to_link)
                except:
                    pass


        # Process intervention_name fields
        if 'intervention_name' in trial_defaults and trial_defaults['intervention_name']:
            for intervention_data in trial_defaults['intervention_name']:
                Intervention.objects.create(
                    trial=trial_obj,
                    intervention_type=intervention_data.get('type', 'Not specified'),
                    intervention_description=intervention_data.get('description', 'No description provided')
                )
                
        # Update trial IDs set
        updated_trial_ids.add(trial_defaults['unique_protocol_id'])

    # Identify and delete obsolete trials
    existing_trial_ids = set(Trial.objects.values_list('unique_protocol_id', flat=True))
    obsolete_trial_ids = existing_trial_ids - updated_trial_ids
    if obsolete_trial_ids:
        Trial.objects.filter(unique_protocol_id__in=obsolete_trial_ids).delete()
        print(f"Deleted {len(obsolete_trial_ids)} obsolete trial records.")

    print(f"Total trials updated or created: {updated_trials}")
    
    # Trigger AI Classification for a sample (to demonstrate log activity)
    print("Starting AI Classification for updated trials (Sample of 3)...")
    llm_logger.info("Starting AI Classification for updated trials (Sample of 3)...")
    
    # Fetch trials that need processing (or just the ones we updated)
    # For now, just pick 3 distinct ones to show it works
    sample_trials = Trial.objects.all().order_by('-study_start_date')[:3]
    
    for trial in sample_trials:
        if trial.eligibility_criteria_generic_description:
            llm_logger.info(f"Processing Trial {trial.unique_protocol_id} with Gemini...")
            response = send_criteria_to_ai_server(trial.unique_protocol_id, trial.eligibility_criteria_generic_description)
            if response:
                trial.eligibility_criteria_inclusion_description = json.dumps(response.get('inclusion', []), cls=DjangoJSONEncoder)
                trial.eligibility_criteria_exclusion_description = json.dumps(response.get('exclusion', []), cls=DjangoJSONEncoder)
                trial.save()
                llm_logger.info(f"Saved AI criteria for {trial.unique_protocol_id}")
            # Rate limit politeness
            time.sleep(5)

    print("AI Sample Processing Complete. Check logs/llm_classification.log")
    llm_logger.info("AI Sample Processing Complete.")

    # After processing, you might want to save data into Excel or perform additional actions
    gene_df = pd.DataFrame(list(Gene.objects.all().values()))
    trial_df = pd.DataFrame(list(Trial.objects.all().values()))
    save_to_xls(gene_df, trial_df)


# Populates the dataframes from 'update_data' function into a three-tab XLS file.
def save_to_xls(gene_df, trial_df):
    # Create a new Excel workbook
    wb = Workbook()
    
    # Create sheets with the specified order
    ws_trial = wb.active  # First sheet is created by default
    ws_trial.title = "Dashboard_trial"
    ws_gene = wb.create_sheet(title="Dashboard_gene")
    ws_trial_genes = wb.create_sheet(title="Dashboard_trial_genes")
    
    # Populate sheets with data frames
    data_frames = [trial_df, gene_df]
    sheets = [ws_trial, ws_gene]
    for df, ws in zip(data_frames, sheets):
        for r in dataframe_to_rows(df, index=False, header=True):
            # Convert lists to strings before appending to Excel
            converted_row = [str(cell) if isinstance(cell, list) else cell for cell in r]
            ws.append(converted_row)
    
    # Add disclosure/citation to the Dashboard_gene sheet
    citation_text = "Citation: ALS/FTD gene data has been obtained from https://ALSoD.ac.uk. Updates to variants for many genes in ALSOD since 2015 have been taken from the supplementary material of Emily McCann et al (2021), which has been a valuable resource. Source: https://alsod.ac.uk/acknowledgements.php"
    ws_gene.append([])  # Add an empty row for spacing
    ws_gene.append([citation_text])  # Add the disclosure statement
    
    # Save the workbook to the specified file path
    file_path = os.path.join(settings.MEDIA_ROOT, 'ALS, FTD Clinical Trial Research Data Tables.xlsx')
    wb.save(file_path)
    print(f"File saved to {file_path}")




# This function creates a column for the URL of each trial record based upon ClinicalTrial.gov's URL schema.
def enhanced_fetch_trial_data():
    # Fetch trial data
    trials_data_df = fetch_trial_data()
    print(f"Type of trials_data_raw before conversion: {type(trials_data_df).__name__}")

    # Initialize new columns with empty lists (conceptually similar to empty JSON arrays)
    # This is done in preperation for OpenAI's interpreting the eligibility_criteria_generic_description field and separating inclusion/exclusion criteria
    trials_data_df['eligibility_criteria_inclusion_description'] = [list() for _ in range(len(trials_data_df))]
    trials_data_df['eligibility_criteria_exclusion_description'] = [list() for _ in range(len(trials_data_df))]

    # Obtain Gene List
    gene_list_df = scrape_alsod_gene_list()
    gene_symbols = gene_list_df['Gene Symbol'].tolist()
    gene_names = gene_list_df['Gene Name'].tolist()  # Assuming this column exists
    gene_name_to_symbol = dict(zip(gene_list_df['Gene Name'], gene_list_df['Gene Symbol']))

    # Define fields to be combined
    fields_to_combine = ['keyword', 'condition', 'study_population', 'brief_title']

    # Generate clinical trial URL
    def clinical_trial_url(nct_id):
        return f"https://clinicaltrials.gov/study/{nct_id}"

    processed_data_df = trials_data_df[fields_to_combine + ['unique_protocol_id']].copy()

    # Combine keywords
    processed_data_df['combined_keywords'] = processed_data_df.apply(
        lambda row: ','.join([','.join(row[field]) if isinstance(row[field], list) else row[field] for field in fields_to_combine]), axis=1)

    # Match genes
    processed_data_df['genes'] = processed_data_df.apply(
        lambda row: match_genes(row['combined_keywords'], gene_symbols, gene_names, gene_name_to_symbol), axis=1)

    # Merge processed data
    trials_data_df = pd.merge(trials_data_df, processed_data_df[['unique_protocol_id', 'genes']], on='unique_protocol_id', how='left')

    # Add clinical_trial_url column
    trials_data_df['clinical_trial_url'] = trials_data_df['nct_id'].apply(clinical_trial_url)

    # Validate and transform choice fields
    trials_data_df['expanded_access'] = trials_data_df['expanded_access'].apply(validate_and_transform_choice_field)
    # Debugging: Print out how many records have expanded_access set to 'TRUE' before the update
    print("Records with expanded_access 'TRUE' before update:", trials_data_df[trials_data_df['expanded_access'] == 'TRUE'].shape[0])
    trials_data_df['fda_regulated_drug'] = trials_data_df['fda_regulated_drug'].apply(validate_and_transform_choice_field)
    trials_data_df['fda_regulated_device'] = trials_data_df['fda_regulated_device'].apply(validate_and_transform_choice_field)
    trials_data_df['eligibility_criteria_healthy_volunteers'] = trials_data_df['eligibility_criteria_healthy_volunteers'].apply(validate_and_transform_choice_field)

    # Update study_phase based on expanded_access being 'TRUE'
    trials_data_df['study_phase'] = trials_data_df.apply(lambda row: 'EAP' if row['expanded_access'] == 'TRUE' else row['study_phase'], axis=1)
    # Verify the update by checking how many records now have study_phase set to 'EAP'
    print("Records with study_phase 'EAP' after update:", trials_data_df[trials_data_df['study_phase'] == 'EAP'].shape[0])
    # Drop the expanded_access field
    trials_data_df.drop('expanded_access', axis=1, inplace=True)

    # Convert Study Phase data
    trials_data_df['study_phase'] = trials_data_df['study_phase'].apply(convert_study_phase)

    return trials_data_df


# Converts study phase data from API.
def convert_study_phase(study_phase_list):
    # If the input is a string, wrap it in a list
    if isinstance(study_phase_list, str):
        study_phase_list = [study_phase_list]

    # Normalize the input list to match the expected format
    normalized_phase_list = [phase.replace('PHASE', 'Phase') for phase in study_phase_list]
    print(f"Normalized study_phase_list: {normalized_phase_list}")  # Debug print

    # Define a mapping from list input to desired output
    phase_mapping = {
        (): 'NA',
        (None,): 'NA',
        ('NA',): 'NA',
        ('EARLY_Phase1',): 'Early Phase 1',
        ('Phase1',): 'Phase 1',
        ('Phase1', 'Phase2'): 'Phase 1/2',
        ('Phase2',): 'Phase 2',
        ('Phase2', 'Phase3'): 'Phase 2/3',
        ('Phase3',): 'Phase 3',
        ('Phase4',): 'Phase 4',
        ('EAP',): 'EAP',  # Ensure 'EAP' is correctly recognized
    }

    # Convert the normalized input list to a tuple for mapping
    study_phase_key = tuple(normalized_phase_list) if normalized_phase_list else ()

    print(f"Converted study_phase_key: {study_phase_key}")  # Debug print

    # Attempt to get the mapped value
    mapped_value = phase_mapping.get(study_phase_key, 'Unknown')
    
    print(f"Mapped value: {mapped_value}")  # Debug print

    return mapped_value



# Data validation for TRUE, FALSE, or BLANK/NULL Choice Fields
def validate_and_transform_choice_field(value):
    """Transform boolean and 'True'/'False' strings to 'TRUE'/'FALSE', allowing blank fields."""
    # Handle boolean values directly
    if value is True:
        return 'TRUE'
    elif value is False:
        return 'FALSE'
    # Handle string values that are 'true'/'false' (case insensitive)
    elif isinstance(value, str):
        lower_value = value.lower()
        if lower_value == 'true':
            return 'TRUE'
        elif lower_value == 'false':
            return 'FALSE'
        elif value.strip() == '':  # Check for a blank string
            return ''
    # Default case for handling any unexpected value types
    return ''


# This function fetches trial data and enhances it by associating gene symbols based on keywords, using the scraped gene list.
# To refine the matching process and avoid false positives, an additional check occurs to ensure that gene symbols
# are recognized as distinct words or part of a larger keyword that correctly represents
# a gene symbol within the context (e.g., at the beginning of a word, followed by a non-alphabetic character, or at the end of a word). 
def match_genes(combined_keywords, gene_symbols, gene_names, gene_name_to_symbol):
    valid_keywords_count = 0
    invalid_keywords_count = 0
    matched_genes = set()  # Use a set to avoid duplicates

    if isinstance(combined_keywords, str) and combined_keywords.strip():
        try:
            # Split the string by comma and strip each keyword
            keywords_list = [keyword.strip() for keyword in combined_keywords.split(',')]

            # Use regex to match gene symbols and gene names more accurately
            for keyword in keywords_list:
                matched_this_round = False
                for gene_symbol in gene_symbols:
                    # Pattern for gene symbol matches
                    pattern_symbol = rf'\b{gene_symbol}\b|\b{gene_symbol}[^A-Za-z0-9_]|[^A-Za-z0-9_]{gene_symbol}\b'
                    if re.search(pattern_symbol, keyword, re.IGNORECASE):
                        matched_genes.add(gene_symbol)
                        valid_keywords_count += 1
                        matched_this_round = True
                        break  # Break if a gene symbol match is found
                
                if not matched_this_round:
                    for gene_name in gene_names:
                        # Pattern for gene name matches, allowing dashes
                        pattern_name = rf'\b{gene_name}\b'
                        if re.search(pattern_name, keyword, re.IGNORECASE):
                            # Add the gene symbol equivalent for the matched gene name
                            matched_genes.add(gene_name_to_symbol[gene_name])
                            valid_keywords_count += 1
                            matched_this_round = True
                            break  # Break if a gene name match is found

                if not matched_this_round:
                    invalid_keywords_count += 1

        except (ValueError, TypeError) as e:
            print(f"Error processing combined keywords: {e}")
            invalid_keywords_count += 1
    else:
        print("Invalid or empty 'combined_keywords' field")
        invalid_keywords_count += 1

    # Print results for debug purposes
    print(f"Total records with valid keyword matches: {valid_keywords_count}")
    print(f"Valid keywords count: {valid_keywords_count}")
    print(f"Invalid/Empty keywords count: {invalid_keywords_count}")
    return json.dumps(list(matched_genes))  # Convert set back to list for JSON serialization


# This function scrapes gene information from Dr. Al-Chalabi's ALSoD.ac.uk HTML table.
# It processes it into a structured format, ready for database insertion or association with trials, while handling text sanitization.
def scrape_alsod_gene_list():
    # Check the last update date from Update_Log table for 'Dashboard_gene' table
    try:
        update_log = Update_Log.objects.get(database_table_name='Dashboard_gene')
        last_update_date = update_log.table_update_date
    except Update_Log.DoesNotExist:
        # If there is no record, consider last update date as None
        last_update_date = None

    # Calculate the difference or set a default to proceed with scraping
    if Gene.objects.count() == 0 or last_update_date is None or (timezone.now().date() - last_update_date.date() >= timedelta(days=30)):
        url = "https://alsod.ac.uk/"
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8"
        }
        
        gene_symbols = []
        gene_names = []
        gene_risk_categories = []

        try:
            print(f"Scraping genes from {url}...")
            response = requests.get(url, headers=headers, timeout=10)
            response.raise_for_status()
            soup = BeautifulSoup(response.text, "html.parser")
            rows = soup.findAll("tr", class_="clickable-row")

            def sanitize(text):
                text = text.replace('"', '')
                text = re.sub(r'\(.*?\)', '', text)
                text = ' '.join(text.split())
                return text

            for row in rows:
                columns = row.findAll("td", class_="assetIDConfig")
                if len(columns) >= 3:
                    gene_symbol = sanitize(columns[1].text)
                    gene_name = sanitize(columns[2].text)
                    gene_risk_category = sanitize(columns[3].text)

                    gene_symbols.append(gene_symbol)
                    gene_names.append(gene_name)
                    gene_risk_categories.append(gene_risk_category)

        except Exception as e:
            print(f"Error scraping ALSoD: {e}")

        # Fallback if scraping failed or returned 0 records (e.g. 403 Forbidden or site change)
        if len(gene_symbols) == 0:
            print("⚠️ Scraping returned 0 genes (likely 403 Forbidden or structure change). Using fallback gene list.")
            fallback_genes = [
                ("SOD1", "Superoxide dismutase 1", "Definitive"),
                ("FUS", "Fused in sarcoma", "Definitive"),
                ("TARDBP", "TAR DNA binding protein", "Definitive"),
                ("C9orf72", "C9orf72-SMCR8 complex subunit", "Definitive"),
                ("TBK1", "TANK binding kinase 1", "Definitive"),
                ("ANG", "Angiogenin", "Definitive"),
                ("OPTN", "Optineurin", "Definitive"),
                ("VCP", "Valosin containing protein", "Definitive"),
                ("UBQLN2", "Ubiquilin 2", "Definitive"),
                ("SQSTM1", "Sequestosome 1", "Definitive")
            ]
            for sym, name, cat in fallback_genes:
                gene_symbols.append(sym)
                gene_names.append(name)
                gene_risk_categories.append(cat)

        if len(gene_symbols) > 0:
            df = pd.DataFrame({
                "Gene Symbol": gene_symbols,
                "Gene Name": gene_names,
                "Gene Risk Category": gene_risk_categories
            })
            print(f"Gene list populated with {len(gene_symbols)} records.")
            
            # Update the Update_Log table
            try:
                Update_Log.objects.update_or_create(
                    database_table_name='Dashboard_gene', 
                    defaults={'table_update_date': timezone.now()}
                )
            except Exception as e:
                print(f"Could not update log: {e}")
            return df

    # If recent update exists OR scraping failed completely, load from DB
    # Query the 'Dashboard_gene' table directly
    genes = Gene.objects.all().values('gene_symbol', 'gene_name', 'gene_risk_category')
    if genes:
        df = pd.DataFrame(list(genes))
        df.columns = ['Gene Symbol', 'Gene Name', 'Gene Risk Category']
        print("Loaded genes from database.")
    else:
        print("Warning: No genes in DB and scraping failed.")
        df = pd.DataFrame(columns=['Gene Symbol', 'Gene Name', 'Gene Risk Category'])

    print("DataFrame shape:", df.shape)
    return df


# This function fetches trial data from ClinicalTrials.gov's API.
# It processes the response, and returns a structured format of trial details.
def fetch_trial_data():
    # ClinicalTrials.gov's API V2 URL.
    base_url = "https://clinicaltrials.gov/api/v2/studies"
    # Conditions to include in search.
    include_conditions = [
        "ALS", "amyotrophic lateral sclerosis", 
        "Amyotrophic Lateral Sclerosis", "Lou Gehrig’s disease", "Lou Gehrigs Disease",
        "Motor Neuron Disease", "motor neuron disease", "MND", "mnd", "frontal lobe dementia", 
        "Frontal Lobe Dementia", "Frontotemporal dementia", "Frontotemporal Dementia",
        "frontotemporal dementia", "FTD", "Behavioral variant Frontotemporal Dementia (bvFTD)",
        "bvFTD", "Frontotemporal Lobar Degeneration (FTLD)", "FTLD", "Frontotemporal Lobar Degeneration"
    ]

    # Constructing the query condition string.
    # Join include conditions with 'OR'
    query_cond = " OR ".join(include_conditions)

    fields = [
        "OrgStudyId", "NCTId", "BriefTitle", "StudyType", "OverallStatus",
        "StatusVerifiedDate", "CompletionDate", "LeadSponsorName",
        "ResponsiblePartyType", "ResponsiblePartyInvestigatorFullName",
        "Condition", "Keyword", "InterventionType", "InterventionDescription",
        "StudyPopulation", "EnrollmentCount", "EnrollmentType", "Phase",
        "StartDate", "StartDateType", "StudyFirstSubmitDate",
        "StudyFirstSubmitQCDate", "HasExpandedAccess", "IsFDARegulatedDrug",
        "IsFDARegulatedDevice", "Location", "PrimaryOutcome", "SecondaryOutcome",
        "OtherOutcome", "EligibilityCriteria", "HealthyVolunteers", "Sex",
        "MinimumAge", "MaximumAge"
    ]
    query_params = {
        "format": "json",
        "query.cond": query_cond,
        "pageSize": 1000,
        "fields": "|".join(fields)
    }
    print("Starting fetch_trial_data..before 'all_study_details'...")
    all_studies_details = []
    print("Starting fetch_trial_data")
    try:
        print("About to make API call")
        while True:
            response = requests.get(base_url, params=query_params)
            print("Finished request.get of base_url")
            print("Response content:", response.content)
            if response.status_code == 200:
                print("response.status_code == 200....proceed to data = response.json")
                data = response.json()

                studies_details = data.get("studies", [])
                all_studies_details.extend(studies_details)
                # Continues to next 1000 page query due to ClinicalTrials.gov's imposed limitations.
                nextPageToken = data.get("nextPageToken")
                if not nextPageToken:
                    break

                query_params["pageToken"] = nextPageToken
            else:
                print(f"Error: Failed to fetch data. Status code: {response.status_code}")
                return {"error": "Failed to fetch data", "status_code": response.status_code}
    except Exception as e:
        print("An error occurred during API call:", e)
        return {"error": str(e)}

    # False-positive conditions from ClinicalTrials.gov needing excluded from dataset.
    # Different trials often times refer to the same condition using slightly different context.
    # I found some trials misspell the conditions; requiring misspelling of exclusion criteria.
    exclude_conditions = [
        # Spinal Muscular Atrophy
        "Spinal Muscular Atrophy", "spinal muscular atrophy", "Spinal Muscular Atrophy (SMA)",
        "SMA", "Type * Spinal Muscular Atrophy", "Type * SMA",
        "Infantile-onset Spinal Muscular Atrophy", "Infantile-onset SMA",
        "Muscular Atrophy, Spinal, Type *", "Muscular Atrophy, Spinal",
        "Type 2 Spinal Muscular Atrophy", "Spinal Muscular Atrophy 1", "Spinal Muscular Atrophy Type II",
        "Spinal Muscular Atrophy Type III", "SMA II", "SMA - Spinal Muscular Atrophy", "Spinal Muscular Atrophy Type I",
        "Chest Deformities", "Spinal Orthosis", "Pulmonary Rehabilitation", "Spinal Muscular Atrophy Type 3",
        "Infantile Spinal Muscular Atrophy, Type I [Werdnig- Hoffman]", "Infantile Spinal Muscular Atrophy",
        "Natural History of Type 1 Spinal Muscular Atrophy (SMA)", "Spinal and Bulbar Muscular Atrophy",
        # Primary Progressive Aphasia
        "Aphasia, Primary Progressive", "Primary Progressive Aphasia", "primary progressive aphasia",
        "PPA", "Logopenic Progressive Aphasia", "logopenic progressive aphasia",
        "Nonfluent Variant Primary Progressive Aphasia (nfvPPA)",
        "Logopenic Variant Primary Progressive Aphasia", "Non-fluent Variant Primary Progressive Aphasia",
        "Semantic Variant Primary Progressive Aphasia",
        # Corticobasal Degeneration and Progressive Supranuclear Palsy
        "Corticobasal Degeneration (CBD)", "Corticobasal Syndrome (CBS)", "Cortical-basal Ganglionic Degeneration (CBGD)",
        "Progressive Supranuclear Palsy (PSP)", "Nonfluent Variant Primary Progressive Aphasia (nfvPPA)",
        "Oligosymptomatic/Variant Progressive Supranuclear Palsy (o/vPSP)",
        "CBD", "CBS", "CBGD", "PSP", "nfvPPA", "oPSP", "vPSP", "o/vPSP",
        "Corticobasal Degeneration", "Corticobasal Syndrome", "Cortocal-basal Ganglionic Degeneration",
        "Progressive Supranuclear Palsy", "Oligosymptomatic Progressive Supranuclear Palsy",
        # Niemann-Pick Disease
        "Niemann-Pick Disease", "Neimann-Pick Disease", "Niemann-Pick Disease, Type C", "Niemann-Pick Disease, Type C*",
        "Niemann-Pick Disease, Type C1", "Niemann-Pick Diseases", "Niemann-Pick Disease Type C*", "Pick Disease of the Brain",
        "Niemann-Pick Type C Disease",
        # Other diseases
        "Chronic Lymphocytic Leukemia", "Hurler Syndrome (MPS I)", "Hurler-Scheie Syndrome", "Hunter Syndrome (MPS II)",
        "Sanfilippo Syndrome (MPS III)", "Krabbe Disease (Globoid Leukodystrophy)", "Metachromatic Leukodystrophy",
        "Adrenoleukodystrophy (ALD and AMN)", "Sandhoff Disease", "Tay Sachs Disease", "Pelizaeus Merzbacher (PMD)",
        "Alpha-mannosidosis", "Juvenile Neuronal Ceroid Lipofuscinosis", "Smith-Lemli-Opitz Syndrome",
        "Creatine Transporter Deficiency", "Mucopolysaccharidosis I", "Mucopolysaccharidosis VI", "Adrenoleukodystrophy",
        "Metachromatic Leukodystrophy", "Wolman Disease", "Krabbe's Disease", "Gaucher's Disease", "Fucosidosis",
        "Batten Disease", "Severe Aplastic Anemia", "Diamond-Blackfan Anemia", "Amegakaryocytic Thrombocytopenia",
        "Myelodysplastic Syndrome", "Acute Myelogenous Leukemia", "Acute Lymphocytic Leukemia", "Lysosomal Acid Lipase Deficiency",
        "Stroke", "Spasticity", "Acid Sphingomyelinase Deficiency", "Gaucher Disease", "ASMD", "Splenomegaly",
        "Colorectal Neoplasms", "Trifluridine and Tipiracil", "Circulating Tumor DNA", "Satisfaction", "Nonalcoholic Steatohepatitis",
        "Prostate Cancer", "Renal Cancer", "Brain Cancer", "Feasibility of Neonatal Screening for Spinal Amyotrophy", "Diplegic Cerebral Palsy",
        "Rare Disorders", "Undiagnosed Disorders", "Disorders of Unknown Prevalence", "Cornelia De Lange Syndrome",
        "Prenatal Benign Hypophosphatasia", "Perinatal Lethal Hypophosphatasia", "Odontohypophosphatasia", "Adult Hypophosphatasia",
        "Childhood-onset Hypophosphatasia", "Infantile Hypophosphatasia", "Hypophosphatasia", "Kabuki Syndrome",
        "Bohring-Opitz Syndrome", "Narcolepsy Without Cataplexy", "Narcolepsy-cataplexy", "Hypersomnolence Disorder",
        "Idiopathic Hypersomnia Without Long Sleep Time", "Idiopathic Hypersomnia With Long Sleep Time", "Idiopathic Hypersomnia",
        "Kleine-Levin Syndrome", "Kawasaki Disease", "Leiomyosarcoma", "Leiomyosarcoma of the Corpus Uteri",
        "Leiomyosarcoma of the Cervix Uteri", "Leiomyosarcoma of Small Intestine", "Acquired Myasthenia Gravis",
        "Addison Disease", "Hyperacusis (Hyperacousis)", "Juvenile Myasthenia Gravis", "Transient Neonatal Myasthenia Gravis",
        "Williams Syndrome", "Lyme Disease", "Myasthenia Gravis", "Marinesco Sjogren Syndrome(Marinesco-Sjogren Syndrome)",
        "Isolated Klippel-Feil Syndrome", "Frasier Syndrome", "Denys-Drash Syndrome", "Beckwith-Wiedemann Syndrome",
        "Emanuel Syndrome", "Isolated Aniridia", "Axenfeld-Rieger Syndrome", "Aniridia-intellectual Disability Syndrome",
        "Aniridia - Renal Agenesis - Psychomotor Retardation", "Aniridia - Ptosis - Intellectual Disability - Familial Obesity",
        "Aniridia - Cerebellar Ataxia - Intellectual Disability", "Aniridia - Absent Patella", "Aniridia",
        "Peters Anomaly - Cataract", "Peters Anomaly", "Potocki-Shaffer Syndrome",
        "Silver-Russell Syndrome Due to Maternal Uniparental Disomy of Chromosome 11",
        "Silver-Russell Syndrome Due to Imprinting Defect of 11p15", "Silver-Russell Syndrome Due to 11p15 Microduplication",
        "Syndromic Aniridia", "WAGR Syndrome", "Wolf-Hirschhorn Syndrome", "4p16.3 Microduplication Syndrome",
        "4p Deletion Syndrome, Non-Wolf-Hirschhorn Syndrome", "Autosomal Recessive Stickler Syndrome",
        "Stickler Syndrome Type *", "Stickler Syndrome", "Mucolipidosis Type 4", "X-linked Spinocerebellar Ataxia Type *",
        "X-linked Intellectual Disability - Ataxia - Apraxia", "Vitamin B12 Deficiency Ataxia", "Toxic Exposure Ataxia",
        "Unclassified Autosomal Dominant Spinocerebellar Ataxia", "Thyroid Antibody Ataxia", "Sporadic Adult-onset Ataxia of Unknown Etiology",
        "Spinocerebellar Ataxia With Oculomotor Anomaly", "Spinocerebellar Ataxia With Epilepsy", "Spinocerebellar Ataxia With Axonal Neuropathy Type *",
        "Spinocerebellar Ataxia Type *", "Spinocerebellar Ataxia - Unknown", "Spinocerebellar Ataxia - Dysmorphism",
        "Non Progressive Epilepsy and/or Ataxia With Myoclonus as a Major Feature", "Spasticity-ataxia-gait Anomalies Syndrome",
        "Spastic Ataxia With Congenital Miosis", "Spastic Ataxia - Corneal Dystrophy", "Spastic Ataxia", "Rare Hereditary Ataxia",
        "Rare Ataxia", "Recessive Mitochondrial Ataxia Syndrome", "Progressive Epilepsy and/or Ataxia With Myoclonus as a Major Feature",
        "Posterior Column Ataxia - Retinitis Pigmentosa", "Post-Stroke Ataxia", "Post-Head Injury Ataxia", "Post Vaccination Ataxia",
        "Polyneuropathy - Hearing Loss - Ataxia - Retinitis Pigmentosa - Cataract", "Muscular Atrophy - Ataxia - Retinitis Pigmentosa - Diabetes Mellitus",
        "Non-hereditary Degenerative Ataxia", "Paroxysmal Dystonic Choreathetosis With Episodic Ataxia and Spasticity",
        "Olivopontocerebellar Atrophy - Deafness", "NARP Syndrome", "Myoclonus - Cerebellar Ataxia - Deafness",
        "Multiple System Atrophy, Parkinsonian Type", "Multiple System Atrophy, Cerebellar Type", "Multiple System Atrophy",
        "Maternally-inherited Leigh Syndrome", "Machado-Joseph Disease Type *", "Leigh Syndrome", "Late-onset Ataxia With Dementia",
        "Infection or Post Infection Ataxia", "GAD Ataxia", "Hereditary Episodic Ataxia", "Gliadin/Gluten Ataxia",
        "Friedreich Ataxia", "Fragile X-associated Tremor/Ataxia Syndrome", "Familial Paroxysmal Ataxia",
        "Exposure to Medications Ataxia", "Episodic Ataxia With Slurred Speech", "Episodic Ataxia Unknown Type",
        "Epilepsy and/or Ataxia With Myoclonus as Major Feature", "Early-onset Spastic Ataxia-neuropathy Syndrome",
        "Early-onset Progressive Neurodegeneration - Blindness - Ataxia - Spasticity",
        "Early-onset Cerebellar Ataxia With Retained Tendon Reflexes", "Early-onset Ataxia With Dementia",
        "Childhood-onset Autosomal Recessive Slowly Progressive Spinocerebellar Ataxia", "Dilated Cardiomyopathy With Ataxia",
        "Cataract - Ataxia - Deafness", "Cerebellar Ataxia, Cayman Type", "Cerebellar Ataxia With Peripheral Neuropathy",
        "Cerebellar Ataxia - Hypogonadism", "Cerebellar Ataxia - Ectodermal Dysplasia",
        "Cerebellar Ataxia - Areflexia - Pes Cavus - Optic Atrophy - Sensorineural Hearing Loss", "Brain Tumor Ataxia",
        "Brachydactyly - Nystagmus - Cerebellar Ataxia", "Benign Paroxysmal Tonic Upgaze of Childhood With Ataxia",
        "Autosomal Recessive Syndromic Cerebellar Ataxia", "Autosomal Recessive Spastic Ataxia With Leukoencephalopathy",
        "Autosomal Recessive Spastic Ataxia of Charlevoix-Saguenay",
        "Autosomal Recessive Spastic Ataxia - Optic Atrophy - Dysarthria", "Autosomal Recessive Spastic Ataxia",
        "Autosomal Recessive Metabolic Cerebellar Ataxia",
        "Autosomal Dominant Spinocerebellar Ataxia Due to Repeat Expansions That do Not Encode Polyglutamine",
        "Autosomal Recessive Ataxia, Beauce Type", "Autosomal Recessive Ataxia Due to Ubiquinone Deficiency",
        "Autosomal Recessive Ataxia Due to PEX10 Deficiency",
        "Autosomal Recessive Degenerative and Progressive Cerebellar Ataxia",
        "Autosomal Recessive Congenital Cerebellar Ataxia Due to MGLUR1 Deficiency",
        "Autosomal Recessive Congenital Cerebellar Ataxia Due to GRID2 Deficiency",
        "Autosomal Recessive Congenital Cerebellar Ataxia",
        "Autosomal Recessive Cerebellar Ataxia-pyramidal Signs-nystagmus-oculomotor Apraxia Syndrome",
        "Autosomal Recessive Cerebellar Ataxia-epilepsy-intellectual Disability Syndrome Due to WWOX Deficiency",
        "Autosomal Recessive Cerebellar Ataxia-epilepsy-intellectual Disability Syndrome Due to TUD Deficiency",
        "Autosomal Recessive Cerebellar Ataxia-epilepsy-intellectual Disability Syndrome Due to KIAA0226 Deficiency",
        "Autosomal Recessive Cerebellar Ataxia-epilepsy-intellectual Disability Syndrome",
        "Autosomal Recessive Cerebellar Ataxia With Late-onset Spasticity", "Autosomal Recessive Cerebellar Ataxia Due to STUB1 Deficiency",
        "Autosomal Recessive Cerebellar Ataxia Due to a DNA Repair Defect",
        "Autosomal Recessive Cerebellar Ataxia - Saccadic Intrusion",
        "Autosomal Recessive Cerebellar Ataxia - Psychomotor Retardation",
        "Autosomal Recessive Cerebellar Ataxia - Blindness - Deafness", "Autosomal Recessive Cerebellar Ataxia",
        "Autosomal Dominant Spinocerebellar Ataxia Due to a Polyglutamine Anomaly",
        "Autosomal Dominant Spinocerebellar Ataxia Due to a Point Mutation",
        "Autosomal Dominant Spinocerebellar Ataxia Due to a Channelopathy",
        "Autosomal Dominant Spastic Ataxia Type *", "Autosomal Dominant Spastic Ataxia", "Autosomal Dominant Optic Atrophy",
        "Ataxia-telangiectasia Variant", "Ataxia-telangiectasia",
        "Autosomal Dominant Cerebellar Ataxia, Deafness and Narcolepsy", "Autosomal Dominant Cerebellar Ataxia Type *",
        "Ataxia-telangiectasia-like Disorder", "Ataxia With Vitamin E Deficiency", "Ataxia With Dementia",
        "Ataxia - Oculomotor Apraxia Type 1", "Ataxia - Other", "Ataxia - Genetic Diagnosis - Unknown", "Acquired Ataxia",
        "Adult-onset Autosomal Recessive Cerebellar Ataxia", "Alcohol Related Ataxia", "Multiple Endocrine Neoplasia",
        "Multiple Endocrine Neoplasia Type *", "Atypical Hemolytic Uremic Syndrome", "Atypical HUS", "Wiedemann-Steiner Syndrome",
        "Breast Implant-Associated Anaplastic Large Cell Lymphoma", "Autoimmune/Inflammatory Syndrome Induced by Adjuvants (ASIA)",
        "Hemophagocytic Lymphohistiocytosis", "Behcet's Disease", "Alagille Syndrome",
        "Inclusion Body Myopathy With Early-onset Paget Disease and Frontotemporal Dementia (IBMPFD)", "Lowe Syndrome", "Pitt Hopkins Syndrome",
        "1p36 Deletion Syndrome", "Jansen Type Metaphyseal Chondrodysplasia", "Cockayne Syndrome", "Chronic Recurrent Multifocal Osteomyelitis",
        "CRMO", "Malan Syndrome", "Hereditary Sensory and Autonomic Neuropathy Type *", "VCP Disease", "Hypnic Jerking", "Sleep Myoclonus",
        "Mollaret Meningitis", "Recurrent Viral Meningitis", "CRB1", "Leber Congenital Amaurosis", "Retinitis Pigmentosa",
        "Rare Retinal Disorder", "KCNMA1-Channelopathy", "Primary Biliary Cirrhosis", "ZMYND11", "Transient Global Amnesia",
        "Glycogen Storage Disease", "Alstrom Syndrome", "White Sutton Syndrome", "DNM1", "EIEE*", "Myhre Syndrome",
        "Recurrent Respiratory Papillomatosis", "Laryngeal Papillomatosis", "Tracheal Papillomatosis", "Refsum Disease",
        "Nicolaides Baraitser Syndrome", "Leukodystrophy", "Tango*", "Cauda Equina Syndrome", "Rare Gastrointestinal Disorders",
        "Achalasia-Addisonian Syndrome", "Achalasia Cardia", "Achalasia Icrocephaly Syndrome", "Anal Fistula",
        "Congenital Sucrase-Isomaltase Deficiency", "Eosinophilic Gastroenteritis", "Idiopathic Gastroparesis", "Hirschsprung Disease",
        "Rare Inflammatory Bowel Disease", "Intestinal Pseudo-Obstruction", "Scleroderma", "Short Bowel Syndrome", "Sacral Agenesis",
        "Sacral Agenesis Syndrome", "Caudal Regression", "Scheuermann Disease", "SMC1A Truncated Mutations (Causing Loss of Gene Function)",
        "Cystinosis", "Juvenile Nephropathic Cystinosis", "Nephropathic Cystinosis", "Kennedy Disease", "Spinal Bulbar Muscular Atrophy",
        "Warburg Micro Syndrome", "Mucolipidoses", "Mitochondrial Diseases", "Mitochondrial Aminoacyl-tRNA Synthetases",
        "Mt-aaRS Disorders", "Hypertrophic Olivary Degeneration", "Non-Ketotic Hyperglycinemia", "Fish Odor Syndrome", "Halitosis",
        "Isolated Congenital Asplenia", "Lambert Eaton (LEMS)", "Biliary Atresia", "STAG1 Gene Mutation", "Coffin Lowry Syndrome",
        "Borjeson-Forssman-Lehman Syndrome", "Blau Syndrome", "Arginase 1 Deficiency", "HSPB8 Myopathy", "Beta-Mannosidosis",
        "TBX4 Syndrome", "DHDDS Gene Mutations", "MAND-MBD5-Associated Neurodevelopmental Disorder", "Constitutional Mismatch Repair Deficiency (CMMRD)",
        "SPATA5 Disorder", "SPATA5L* Related Disorder", "Kennedy's Disease", "Sialorrhea", "Fibromyalgia", "Fertility Issues", "Asmd, Visceral Type",
        "Sphingomyelin Lipidosis", "Cholangiocarcinoma", "Stage III Gallbladder Cancer AJCC v7", "Stage IIIA Gallbladder Cancer AJCC v7",
        "Stage IIIB Gallbladder Cancer AJCC v7", "Stage IV Gallbladder Cancer AJCC v7", "Stage IVA Gallbladder Cancer AJCC v7",
        "Stage IVB Gallbladder Cancer AJCC v7", "Hemiplegic Cerebral Palsy", "Tetraplegia", "Psychiatric Adults Patients", "Alzheimer Disease", "Alzheimer's Disease",
        "Postpoliomyelitis", "Duchenne Muscular Dystrophy", "Inherited Metabolic Diseases", "Lysosomal Storage Disorders", "Peroxisomal Storage Diseases",
        "Inborn Errors of Metabolism", "Mucopolysaccharidosis", "Spinal Cord Injuries", "Death, Sudden, Cardiac", "Out-Of-Hospital Cardiac Arrest",
        "Ventricular Fibrillation", "Cardiopulmonary Arrest With Successful Resuscitation", "Insomnia", "Depression", "Distal Hereditary Motor Neuropathy, Type II",
        "Distal Hereditary Motor Neuropathy, Type V", "Distal Hereditary Motor Neuronopathy Type I", "Distal Hereditary Motor Neuronopathy Type VI",
        "Cerebral Palsy", "Hirayama Disease", "Osteoporosis","Poliomyelitis", "Postpoliomyelitis Syndrome",
        "Children With Spastic Diplegia, Between the Ages of 2 to 10 Years", "Gross Motor Function Classification System (GMFCS) Level I,II and III",
        "Inclusion Body Myopathy With Early-onset Paget Disease and Frontotemporal Dementia", "Paget Disease of Bone",
        "Myopathy", "Pompe Disease (Late-onset)", "Inclusion Body Myositis, Sporadic", "Facioscapulohumeral Muscular Dystrophy 1",
        "Myotonic Dystrophy Type 1 (DM1)", "Myotonic Dystrophy Type 2", "IGF-1 Deficiency" "Stage III Gallbladder Cancer AJCC v7",
        "Stage IIIA Gallbladder Cancer AJCC v7", "Stage IIIB Gallbladder Cancer AJCC v7", "Stage IV Gallbladder Cancer AJCC v7",
        "Stage IVA Gallbladder Cancer AJCC v7", "Stage IVB Gallbladder Cancer AJCC v7", "Hemiplegic Cerebral Palsy", "Tetraplegia",
        "Psychiatric Adults Patients", "Advanced solid tumors", "IGF-1 Deficiency", "Fabry Disease", "Lysosomal Storage Diseases"
    ]

    # Post-processing of fetched API data to apply exclusion criteria:
    # This approach inadvertently filtered out studies that have both relevant and irrelevant conditions. I'm going to try changing it so that, 
    # if a study is included in the dataset, at least one of its conditions listed is not in the exclusion criteria,
    # AND if the condition is a 'fuzzy' match to the inclusion conditions, then the record remains in the dataset.
    # Otherwise, the record is excluded.
    exclude_conditions_set = {cond.lower() for cond in exclude_conditions}
    filtered_studies_details = []
    for study in all_studies_details:
        # Extract and lowercase the conditions for the study
        study_conditions = [cond.lower() for cond in study.get('protocolSection', {}).get('conditionsModule', {}).get('conditions', [])]
        
        # Filter out studies with conditions in the exclude list and then check for fuzzy matches
        for cond in study_conditions:
            if cond not in exclude_conditions_set:
                # Find the best match from included_conditions and its score
                match, score = process.extractOne(cond, include_conditions)
                
                # Define a threshold for considering a match good enough (e.g., 80 out of 100)
                # I am using a more forgiving threshold due to the explicit exclusion criteria defined
                if score >= 75:  # Adjust the threshold as needed
                    filtered_studies_details.append(study)
                    break  # If a match is found, no need to check other conditions for this study

    # Convert to DataFrame for further processing
    df_studies = pd.json_normalize(filtered_studies_details)

    print("After conversion to DataFrame, headers:", df_studies.columns.tolist())
    if not df_studies.empty:
        print("Sample record after conversion to DataFrame:", df_studies.iloc[0])
        print("DataFrame structure:")
        print(df_studies)

    # Print statement for column remapping
    print("Before renaming, headers:", df_studies.columns.tolist())

    # Replace 'nan' strings with actual NaN values
    df_studies.replace('nan', np.nan, inplace=True)

    # Fill missing values with an empty string
    df_studies.fillna('', inplace=True)

    # Renaming API column output to match the database model's field names
    column_mappings = {
        "protocolSection.identificationModule.orgStudyIdInfo.id": "unique_protocol_id", # Primary Key from models.py
        "protocolSection.identificationModule.nctId": "nct_id",
        "protocolSection.identificationModule.briefTitle": "brief_title",
        "protocolSection.designModule.studyType": "study_type",
        "protocolSection.designModule.phases": "study_phase",
        "protocolSection.statusModule.overallStatus": "overall_status",
        "protocolSection.statusModule.studyFirstSubmitDate": "study_submitance_date",
        "protocolSection.statusModule.studyFirstSubmitQcDate": "study_submitance_date_qc",
        "protocolSection.statusModule.startDateStruct.date": "study_start_date",
        "protocolSection.statusModule.startDateStruct.type": "study_start_date_type",
        "protocolSection.statusModule.statusVerifiedDate": "status_verified_date",
        "protocolSection.statusModule.completionDateStruct.date": "completion_date",
        "protocolSection.sponsorCollaboratorsModule.leadSponsor.name": "lead_sponsor_name",
        "protocolSection.sponsorCollaboratorsModule.responsibleParty.type": "responsible_party_type",
        "protocolSection.sponsorCollaboratorsModule.responsibleParty.investigatorFullName": "responsible_party_investigator_full_name",
        "protocolSection.conditionsModule.conditions": "condition",
        "protocolSection.sponsorCollaboratorsModule.collaborators": "collaborators",
        "protocolSection.conditionsModule.keywords": "keyword",
        "protocolSection.armsInterventionsModule.interventions": "intervention_types",
        "protocolSection.armsInterventionsModule.interventions": "intervention_name",
        "protocolSection.eligibilityModule.studyPopulation": "study_population",
        "protocolSection.designModule.enrollmentInfo.count": "enrollment_count",
        "protocolSection.designModule.enrollmentInfo.type": "enrollment_type",
        "protocolSection.statusModule.expandedAccessInfo.hasExpandedAccess": "expanded_access",
        "protocolSection.oversightModule.isFdaRegulatedDrug": "fda_regulated_drug",
        "protocolSection.contactsLocationsModule.locations": "study_location",
        "protocolSection.oversightModule.isFdaRegulatedDevice": "fda_regulated_device",
        "protocolSection.outcomesModule.primaryOutcomes": "primary_outcomes",
        "protocolSection.outcomesModule.secondaryOutcomes": "secondary_outcomes",
        "protocolSection.outcomesModule.otherOutcomes": "other_outcomes",
        "protocolSection.eligibilityModule.eligibilityCriteria": "eligibility_criteria_generic_description",
        "protocolSection.eligibilityModule.healthyVolunteers": "eligibility_criteria_healthy_volunteers",
        "protocolSection.eligibilityModule.sex": "eligibility_criteria_sex",
        "protocolSection.eligibilityModule.minimumAge": "eligibility_criteria_min_age_years",
        "protocolSection.eligibilityModule.maximumAge": "eligibility_criteria_max_age_years",
    }

    df_studies.rename(columns=column_mappings, inplace=True)
    print("After renaming, headers:", df_studies.columns.tolist())
    if not df_studies.empty:
        print("Sample record after renaming:", df_studies.iloc[0])

    # Additional print statements for debugging
    if 'nct_id' in df_studies.columns:
        print("First ten records' 'nct_id' field:")
        print(df_studies['nct_id'].head(10))
    else:
        print("The 'nct_id' column does not exist in the DataFrame.")

    # Update the Update_Log table with the current date as the last update date for 'Dashboard_trial'
        Update_Log.objects.update_or_create(
            database_table_name='Dashboard_trial', 
            defaults={'table_update_date': timezone.now().date()}
        )

    # Print the first ten records' 'nct_id' field if it exists
    if 'nct_id' in df_studies.columns:
        print("First ten records' 'nct_id' field:")
        print(df_studies['nct_id'].head(50))
    else:
        print("The 'nct_id' column does not exist in the DataFrame.")

    # Directly return the DataFrame
    return df_studies

# AI Utility Functions Referenced within views.py File
def parse_criteria_from_response(response_text):
    # Split the response into two parts: inclusion and exclusion criteria
    parts = response_text.split("**Exclusion Criteria:**")
    inclusion_text = parts[0].split("**Inclusion Criteria:**")[1] if "**Inclusion Criteria:**" in parts[0] else ""
    exclusion_text = parts[1] if len(parts) > 1 else ""

    # Extract list items from each part
    inclusion_criteria = extract_list_items(inclusion_text)
    exclusion_criteria = extract_list_items(exclusion_text)

    return inclusion_criteria, exclusion_criteria

def extract_list_items(text):
    # Define a regex pattern that matches lines starting with "-", "*", or a number followed by "."
    pattern = re.compile(r'^(\d+\.\s*|\-\s*|\*\s*)')
    list_items = []

    for line in text.split('\n'):
        if pattern.match(line.strip()):
            # Remove the list item prefix (numbers, dashes, or asterisks) and any following whitespace
            item = pattern.sub('', line).strip()
            list_items.append(item)
    
    return json.dumps(list_items, ensure_ascii=False)

# JSON Schemas for Local LLM
CRITERIA_SCHEMA = {
    "name": "clinical_criteria",
    "strict": True,
    "schema": {
        "type": "object",
        "properties": {
            "inclusion": {
                "type": "array",
                "items": {"type": "string"},
                "description": "List of inclusion criteria"
            },
            "exclusion": {
                "type": "array",
                "items": {"type": "string"},
                "description": "List of exclusion criteria"
            }
        },
        "required": ["inclusion", "exclusion"],
        "additionalProperties": False
    }
}

FACILITY_SCHEMA = {
    "name": "facility_location",
    "strict": True,
    "schema": {
        "type": "object",
        "properties": {
            "city": {"type": "string"},
            "zip_code": {"type": "string"},
            "country": {"type": "string"},
            "latitude": {"type": "number"},
            "longitude": {"type": "number"}
        },
        "required": ["city", "zip_code", "country", "latitude", "longitude"],
        "additionalProperties": False
    }
}

def send_criteria_to_ai_server(unique_protocol_id, eligibility_criteria):
    llm_logger.info(f"Preparing to send data for unique_protocol_id: {unique_protocol_id}")
    
    api_key = os.environ.get("LLM_API_KEY")
    base_url = os.environ.get("LLM_BASE_URL")
    model_name = os.environ.get("LLM_MODEL")

    # Priority: Check for Local LLM overrides
    local_url = os.environ.get("LOCAL_LLM_URL")
    if local_url:
        llm_logger.info(f"Using Local LLM at {local_url}")
        base_url = local_url
        api_key = "lm-studio" # Dummy key for local
        model_name = os.environ.get("LOCAL_LLM_MODEL", "local-model")

    if not api_key:
        llm_logger.error("❌ LLM_API_KEY not found in environment variables.")
        return {"inclusion": [], "exclusion": []}

    if not base_url:
        llm_logger.error("❌ LLM_BASE_URL not found in environment variables.")
        return {"inclusion": [], "exclusion": []}

    # Use generic OpenAI-compatible endpoint
    client = OpenAI(
        base_url=base_url,
        api_key=api_key
    )
    criteria_responses = {"inclusion": [], "exclusion": []}

    # Enhanced Prompt for JSON Extraction
    # Enhanced Prompt for JSON Extraction
    prompt = (
        "You are an expert Clinical Data Curator. Your task is to extract, clean, and structure eligibility criteria from the provided text.\n"
        "The input text contains Inclusion and Exclusion criteria, often mixed or poorly formatted.\n\n"
        "Instructions:\n"
        "1. **Analyze** the text to clearly identify the boundary between 'Inclusion Criteria' and 'Exclusion Criteria'.\n"
        "2. **Extract** each distinct criterion as a separate string in the respective array.\n"
        "3. **Refine**: Process each criterion:\n"
        "   - Remove bullet points, numbering (e.g., '1.', '-'), and extra whitespace.\n"
        "   - Split long, complex paragraphs into individual, atomic requirements.\n"
        "   - Simplify wording where possible without losing medical precision.\n"
        "4. **Format**: Return ONLY valid JSON with keys 'inclusion' and 'exclusion'.\n"
        "5. If a section is completely missing, return an empty array for that key.\n\n"
        f"Input Text:\n{eligibility_criteria}\n\n"
        "Structured JSON Response:"
    )

    max_retries = 3
    base_delay = 20 # seconds

    # Determine response format based on backend
    resp_format = {"type": "json_object"}
    if os.environ.get("LOCAL_LLM_URL"):
        resp_format = {
            "type": "json_schema",
            "json_schema": CRITERIA_SCHEMA
        }

    for attempt in range(max_retries):
        try:
            completion = client.chat.completions.create(
                model=model_name,
                messages=[
                    {"role": "system", "content": "You are a helpful assistant that extracts clinical trial criteria as JSON."},
                    {"role": "user", "content": prompt}
                ],
                response_format=resp_format,
                temperature=0.1,
                max_tokens=4096
            )

            if completion.choices and len(completion.choices) > 0:
                response_text = completion.choices[0].message.content
                llm_logger.info(f"Extracted JSON for {unique_protocol_id} (snippet): {response_text[:100]}...")
                
                try:
                    # Clean potential markdown formatting often returned by local models
                    clean_json_text = response_text.replace("```json", "").replace("```", "").strip()
                    
                    # Directly parse the JSON response
                    data = json.loads(clean_json_text)
                    
                    # Ensure the keys exist, defaulting to empty list if not
                    criteria_responses["inclusion"] = data.get("inclusion", [])
                    criteria_responses["exclusion"] = data.get("exclusion", [])
                    
                    # Double check that they are lists, in case the LLM returned strings
                    if isinstance(criteria_responses["inclusion"], str):
                        criteria_responses["inclusion"] = [criteria_responses["inclusion"]]
                    if isinstance(criteria_responses["exclusion"], str):
                        criteria_responses["exclusion"] = [criteria_responses["exclusion"]]
                    
                    return criteria_responses # Success

                except json.JSONDecodeError as json_err:
                     llm_logger.error(f"Failed to parse JSON response for {unique_protocol_id}: {json_err}")
                     llm_logger.error(f"Raw Response: {response_text}")
                     return {"inclusion": [], "exclusion": []}

            else:
                llm_logger.warning(f"No choices found in the response for {unique_protocol_id}.")
                return {"inclusion": [], "exclusion": []}

        except Exception as e:
            error_str = str(e)
            if "429" in error_str or "RESOURCE_EXHAUSTED" in error_str:
                wait_time = base_delay * (attempt + 1)
                llm_logger.warning(f"Rate limited (429) for {unique_protocol_id}. Retrying in {wait_time}s... (Attempt {attempt+1}/{max_retries})")
                time.sleep(wait_time)
            else:
                llm_logger.error(f"Error processing criteria for {unique_protocol_id}: {e}")
                break # Non-retriable error

    return criteria_responses


# Healey Platform Trial Scrape Function
def scrape_healey_platform_trial():
    url = "https://www.massgeneral.org/neurology/als/research/platform-trial-sites"
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3'
    }

    response = requests.get(url, headers=headers)
    soup = BeautifulSoup(response.content, 'html.parser')
    table = soup.find('table')
    rows = table.find_all('tr')[1:]  # Skip the header row

    with transaction.atomic():  # Use a transaction to ensure data integrity
        for row in rows:
            columns = row.find_all('td')
            if len(columns) < 4:
                continue  # Skip rows that do not have enough columns

            trial_data = {
                'facility': columns[0].text.strip(),
                'state': columns[1].text.strip(),
                'enrollment_status': columns[2].text.strip(),
                'trial_contact_info': []
            }

            contacts = columns[3].find_all('a', class_='none')
            for contact in contacts:
                href = contact.get('href', '')
                text = contact.text.strip()
                contact_entry = {}
                if 'mailto:' in href:
                    email = href.split('mailto:')[1]
                    contact_entry = {'trial_contact_name': text, 'trial_contact_email': email}
                elif href.startswith('/'):
                    phone_number = text
                    if trial_data['trial_contact_info']:
                        trial_data['trial_contact_info'][-1]['trial_contact_phone_number'] = phone_number
                    continue
                if contact_entry:  # Ensure we don't add empty entries
                    trial_data['trial_contact_info'].append(contact_entry)

            # Save directly to the database
            HealeyTrial.objects.update_or_create(
            facility=trial_data['facility'],
            state=trial_data['state'],  # Adding another field to enforce uniqueness
            defaults=trial_data
        )

def enrich_facility_data(facility, state):
    llm_logger.info(f"Preparing to enrich data for facility: {facility} in {state}")

    api_key = os.environ.get("LLM_API_KEY")
    base_url = os.environ.get("LLM_BASE_URL")
    model_name = os.environ.get("LLM_MODEL")

    # Priority: Check for Local LLM overrides
    local_url = os.environ.get("LOCAL_LLM_URL")
    if local_url:
        llm_logger.info(f"Using Local LLM at {local_url}")
        base_url = local_url
        api_key = "lm-studio"
        model_name = os.environ.get("LOCAL_LLM_MODEL", "local-model")

    if not api_key:
        llm_logger.error("❌ LLM_API_KEY not found. Skipping facility enrichment.")
        return {}

    if not base_url:
        llm_logger.error("❌ LLM_BASE_URL not found. Skipping facility enrichment.")
        return {}

    # Use generic OpenAI-compatible endpoint
    client = OpenAI(
        base_url=base_url,
        api_key=api_key
    )

    prompt = (
        f"I have a medical facility named '{facility}' in the state of '{state}'.\n"
        "I need its location details.\n"
        "Output valid JSON with keys: city, zip_code, country, latitude, longitude.\n"
        "Example: {\"city\": \"Boston\", \"zip_code\": \"02114\", \"country\": \"USA\", \"latitude\": 42.36, \"longitude\": -71.06}\n"
        "JSON Response:"
    )

    resp_format = {"type": "json_object"}
    if os.environ.get("LOCAL_LLM_URL"):
         resp_format = {
            "type": "json_schema",
            "json_schema": FACILITY_SCHEMA
        }

    try:
        completion = client.chat.completions.create(
            model=model_name,
            messages=[
                {"role": "system", "content": "You are a helpful location assistant. Output valid JSON only."},
                {"role": "user", "content": prompt}
            ],
            response_format=resp_format,
            temperature=0.1,
            max_tokens=1024
        )

        if completion.choices and len(completion.choices) > 0:
            response_text = completion.choices[0].message.content
            llm_logger.info(f"Enriched Data for {facility} (snippet): {response_text}")
            
            try:
                data = json.loads(response_text)
                return data
            except json.JSONDecodeError as json_err:
                 llm_logger.error(f"Failed to parse JSON for {facility}: {json_err}")
                 return {}
        else:
            return {}

    except Exception as e:
        llm_logger.error(f"Error enriching facility {facility}: {e}")
        return {}


def process_healey_trial_data():
    # Scrape data
    scraped_data = scrape_healey_platform_trial()
    for item in scraped_data:
        facility = item['facility']
        state = item['state']
        overall_status = item['overall_status']  # Extracting overall status
        contacts = item['trial_contact_info']  # Extracting trial contact info

        # Save to database
        for contact in contacts:
            trial_contact_name = contact.get('trial_contact_name', '')
            trial_contact_email = contact.get('trial_contact_email', '')
            trial_contact_phone_number = contact.get('trial_contact_phone_number', '')
            HealeyTrial.objects.create(
                facility=facility,
                state=state,
                overall_status=overall_status,
                trial_contact_name=trial_contact_name,
                trial_contact_email=trial_contact_email,
                trial_contact_phone_number=trial_contact_phone_number
            )
